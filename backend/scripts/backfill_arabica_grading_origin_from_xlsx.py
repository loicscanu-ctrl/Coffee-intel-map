"""Backfill historical Arabica grading-by-origin from the synthesis workbook.

The ICE daily XLS only carries the origin × port grading matrix from June 2026
on, and snapshots captured before the matrix parser shipped kept just the
scalar passed/failed totals. The user's master workbook (e.g. Synthesis NY.xlsx,
sheet `6_ny_grading`) holds the full per-(date, port, origin, status) Arabica
grading history — so we can recover the whole back-history from it at once
instead of waiting for, or re-fetching, ICE dailies.

This is a SURGICAL patch: it only adds `passed_by_origin` / `failed_by_origin`
(the live-scraper shape: {origin: {by_port, group, total}}) to snapshots that
already exist in the JSON and are still missing the breakdown. It never
rewrites other fields, never adds/removes snapshots, and validates each day's
re-derived totals against the stored scalars before writing.

Accepts either:
  • the .xlsx workbook (openpyxl; sheet auto-resolved — 6_ny_grading /
    6_ny_gradings / any "*ny*grad*" sheet that isn't pending/queue), or
  • a .csv export of that sheet (columns matched by header name: a date column,
    a port-code column, a status column [pass/fail], a bags column, an origin
    column). CSV is the easy path when the workbook can't be reached directly —
    drop it in the repo and point this script at it.

Usage (run where the file is on disk / in the repo):
    python -m scripts.backfill_arabica_grading_origin_from_xlsx "Synthesis NY.xlsx"
    python -m scripts.backfill_arabica_grading_origin_from_xlsx grading.csv --write
    # Windows xlsx path example:
    #   "C:\\Users\\Loic Scanu\\OneDrive - Tuan Loc Commodities\\Research\\Exchange economics\\Certified stocks\\Synthesis NY.xlsx"

Options:
    --json PATH          target JSON (default: frontend/public/data/certified_stocks_arabica.json)
    --sheet NAME         force a specific sheet name (xlsx only)
    --write              persist the patch
    --allow-mismatch     write workbook detail even when a day's total differs
                         from the stored scalar (default: skip + report)
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from datetime import date
from pathlib import Path

from scraper.sources.ice_arabica_groups import group_of

DEFAULT_JSON = Path("frontend/public/data/certified_stocks_arabica.json")
SHEET_CANDIDATES = ("6_ny_grading", "6_ny_gradings", "6_NY_Grading")


# ── small coercers (mirror import_synthesis_xlsx) ─────────────────────────────

def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(round(float(str(v).replace(",", ""))))
    except (TypeError, ValueError):
        return 0


def _str(v) -> str:
    return "" if v is None else str(v).strip()


def _kind(status: str) -> str | None:
    s = status.lower()
    return "passed" if s.startswith("pass") else "failed" if s.startswith("fail") else None


def _is_action(snap: dict) -> bool:
    return (snap.get("passed_today_bags") or 0) > 0 or (snap.get("failed_today_bags") or 0) > 0


def _has_detail(snap: dict) -> bool:
    return bool(snap.get("passed_by_origin") or snap.get("failed_by_origin"))


# ── row aggregation → {date: {by_port_origin_status, passed, failed}} ─────────

def _new_day():
    return {"passed": 0, "failed": 0,
            "by_port_origin_status": defaultdict(lambda: defaultdict(lambda: defaultdict(int)))}


def _aggregate(rows, since: date) -> dict:
    """rows = iterable of (date, port_code, status, bags, origin)."""
    per_date: dict[date, dict] = defaultdict(_new_day)
    n = 0
    for d, port, status, bags, origin in rows:
        if d is None or d < since:
            continue
        kind = _kind(status)
        if kind is None or bags <= 0 or not port:
            continue
        # Drop aggregate/total pseudo-ports.
        if port.lower() in {"total", "totals", "all", "grand total"}:
            continue
        per_date[d][kind] += bags
        if origin:
            per_date[d]["by_port_origin_status"][port][origin][kind] += bags
        n += 1
    print(f"    parsed {n} grading rows across {len(per_date)} dates.")
    return per_date


def _rows_from_xlsx(path: Path, sheet_name: str | None):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = sheet_name
    if name is None:
        for cand in SHEET_CANDIDATES:
            if cand in wb.sheetnames:
                name = cand
                break
    if name is None:  # fuzzy: an NY grading sheet that isn't pending/queue
        for s in wb.sheetnames:
            low = s.lower()
            if "grad" in low and "ny" in low and "pend" not in low and "queue" not in low:
                name = s
                break
    if name is None or name not in wb.sheetnames:
        raise SystemExit(f"could not find a grading sheet (looked for {SHEET_CANDIDATES}; "
                         f"workbook has: {wb.sheetnames})")
    print(f"    using sheet '{name}'")
    sheet = wb[name]
    # Layout (import_synthesis sheet 6): date | port_code | port_name | status | bags | origin
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 6:
            continue
        yield (_to_date(r[0]), _str(r[1]), _str(r[3]), _int(r[4]), _str(r[5]))


def _rows_from_delimited(path: Path):
    """Read a .csv / .tsv / .txt export of the grading sheet. Delimiter is
    auto-detected (tab vs comma vs semicolon). Columns are matched by header
    name, so column order doesn't matter. Rows with an empty origin (older
    history, before ICE published origins) still feed the day's pass/fail
    totals but carry no origin — handled downstream by leaving the snapshot's
    scalar untouched."""
    raw = path.read_text(encoding="utf-8-sig")
    first = raw.split("\n", 1)[0]
    delim = "\t" if "\t" in first else ";" if (";" in first and "," not in first) else ","
    reader = csv.reader(raw.splitlines(), delimiter=delim)
    header = next(reader, None)
    if not header:
        raise SystemExit("empty file")
    low = [h.strip().lower() for h in header]

    def find(*keys, avoid=()):
        for i, h in enumerate(low):
            if any(k in h for k in keys) and not any(a in h for a in avoid):
                return i
        return None

    i_date = find("date")
    i_port = find("port_code", "port code") or find("port", avoid=("name",)) or find("port")
    i_status = find("status", "pass", "result")
    i_bags = find("bag", "lot", "qty", "quantity", "volume")
    i_origin = find("origin", "growth", "country")
    missing = [n for n, i in (("date", i_date), ("port", i_port), ("status", i_status),
                              ("bags", i_bags), ("origin", i_origin)) if i is None]
    if missing:
        raise SystemExit(f"file missing column(s) {missing}; header was: {header}")
    print(f"    delimiter={delim!r}; columns → date={header[i_date]}, port={header[i_port]}, "
          f"status={header[i_status]}, bags={header[i_bags]}, origin={header[i_origin]}")
    for row in reader:
        if not row or len(row) <= max(i_date, i_port, i_status, i_bags, i_origin):
            continue
        yield (_to_date(row[i_date]), _str(row[i_port]), _str(row[i_status]),
               _int(row[i_bags]), _str(row[i_origin]))


def _to_scraper_shape(by_port_origin_status: dict) -> tuple[dict, dict]:
    """{port: {origin: {passed, failed}}} → two {origin: {by_port, group, total}}."""
    passed: dict[str, dict] = {}
    failed: dict[str, dict] = {}
    for port, by_origin in by_port_origin_status.items():
        for origin, st in by_origin.items():
            for kind, store in (("passed", passed), ("failed", failed)):
                bags = int(st.get(kind, 0) or 0)
                if bags <= 0:
                    continue
                e = store.setdefault(origin, {"by_port": {}, "group": group_of(origin) or "Unknown", "total": 0})
                e["by_port"][port] = e["by_port"].get(port, 0) + bags
                e["total"] += bags
    return passed, failed


def backfill(src: Path, json_path: Path, *, sheet: str | None, write: bool, allow_mismatch: bool) -> dict:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    snaps = data.get("snapshots") or []
    by_date = {s["date"][:10]: s for s in snaps}
    earliest = min(by_date) if by_date else "1900-01-01"
    since = date(*(int(x) for x in earliest.split("-")))

    rows = (_rows_from_delimited(src) if src.suffix.lower() in {".csv", ".tsv", ".txt"}
            else _rows_from_xlsx(src, sheet))
    per_date = _aggregate(rows, since)

    summary = {
        "action_days_in_json": sum(1 for s in snaps if _is_action(s)),
        "already_detailed": sum(1 for s in snaps if _is_action(s) and _has_detail(s)),
        "source_dates": len(per_date),
        "recovered": 0, "not_in_json": 0, "skipped_have_detail": 0, "mismatch": 0,
    }

    for d, rec in sorted(per_date.items()):
        iso = d.isoformat()
        snap = by_date.get(iso)
        if snap is None:
            summary["not_in_json"] += 1
            continue
        if _has_detail(snap):
            summary["skipped_have_detail"] += 1
            continue
        passed_bo, failed_bo = _to_scraper_shape(rec["by_port_origin_status"])
        if not passed_bo and not failed_bo:
            continue
        wb_passed = sum(o["total"] for o in passed_bo.values())
        wb_failed = sum(o["total"] for o in failed_bo.values())
        p_ok = wb_passed == (snap.get("passed_today_bags") or 0)
        f_ok = wb_failed == (snap.get("failed_today_bags") or 0)
        if not (p_ok and f_ok) and not allow_mismatch:
            print(f"    {iso} MISMATCH: source P/F={wb_passed}/{wb_failed} "
                  f"vs stored {snap.get('passed_today_bags')}/{snap.get('failed_today_bags')} — skipped")
            summary["mismatch"] += 1
            continue
        if passed_bo:
            snap["passed_by_origin"] = passed_bo
        if failed_bo:
            snap["failed_by_origin"] = failed_bo
        summary["recovered"] += 1
        flag = "" if (p_ok and f_ok) else "  (mismatch, forced)"
        print(f"    {iso} ✓ {len(passed_bo)} passed-origins, {len(failed_bo)} failed-origins{flag}")

    if write and summary["recovered"]:
        json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        print(f"\nWrote {summary['recovered']} backfilled day(s) to {json_path}")
    elif not write:
        print("\n(dry-run; pass --write to persist)")
    return summary


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("source", type=Path,
                    help="the workbook (.xlsx) or a delimited export of the grading sheet (.csv/.tsv/.txt)")
    ap.add_argument("--json", type=Path, default=DEFAULT_JSON, dest="json_path")
    ap.add_argument("--sheet", default=None, help="force a sheet name (xlsx only)")
    ap.add_argument("--write", action="store_true", help="patch the JSON in place")
    ap.add_argument("--allow-mismatch", action="store_true",
                    help="write source detail even when a day's total differs from the stored scalar")
    args = ap.parse_args()
    summary = backfill(args.source, args.json_path, sheet=args.sheet,
                       write=args.write, allow_mismatch=args.allow_mismatch)
    print(f"\n=== grading backfill · {args.source.name} → {args.json_path} ===")
    for k, v in summary.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
