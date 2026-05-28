#!/usr/bin/env python3
"""import_synthesis_xlsx.py — load an aggregated ICE-data synthesis workbook
into the certified_stocks_*.json shape the frontend reads.

Built specifically for the 13-sheet workbook the user maintains in OneDrive
(Data_ICE_Exchange.xlsx). Each sheet maps to a portion of the JSON output —
see the SHEET_MAP comment block below. The script:

  • Reads each sheet via openpyxl read-only mode (streaming, no full load).
  • Trims to the last DAILY_KEEP_DAYS (default 365) of daily data and the
    last MONTHLY_KEEP_MONTHS (default 24) of monthly data — keeps the JSON
    a sane size (~MB, not tens of MB) while preserving drill-down for any
    period the panel actually shows.
  • Reuses the orchestrator's merge helpers so prior content from the live
    scraper isn't lost.

Run:
    python backend/scripts/import_synthesis_xlsx.py <path-to-xlsx>
    python backend/scripts/import_synthesis_xlsx.py <path> --no-write
    python backend/scripts/import_synthesis_xlsx.py <path> --days 730   # widen daily window
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scraper.sources.ice_arabica_groups import group_of
from scraper.sources.ice_certified_stocks.orchestrate import (
    _merge_arabica, _merge_robusta, OUT_DIR,
)

# ── Tunables ──────────────────────────────────────────────────────────────────
DAILY_KEEP_DAYS    = 365     # last N calendar days kept in daily snapshots
MONTHLY_KEEP_MONTHS = 24     # last N months kept (age allowance + iss/recv monthly)

# ── SHEET_MAP ─────────────────────────────────────────────────────────────────
# 1_ld_tender             → robusta recent_activity.tenders (unpivoted)
# 2_ld_gradings           → robusta recent_activity.gradings (entries)
# 3_age_allowance         → robusta monthly.age_allowance (buckets)
# 4_ld_issuer_receiver    → robusta recent_activity.iss_recv_daily (members)
# 5_ny_pend_gradings      → arabica snapshots[].sections.pending_grading
# 6_ny_gradings           → arabica snapshots[].passed/failed_today_bags + by_port
# 7_ny                    → arabica snapshots[].sections.total_certified
# 8a_ny_issuers           → arabica issuers (NEW field: latest_detail.issuers)
# 8b_ny_stoppers          → arabica stoppers (NEW field: latest_detail.stoppers)
# 9_ld                    → robusta snapshots[].by_port_lots + latest_detail.stock_report
# 12_ny_age               → arabica age detail (NEW field: latest_detail.age_detail)
# 13_ld_gradings_queue    → robusta recent_activity.grading_overview


# ── Date helpers ──────────────────────────────────────────────────────────────

def _to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try: return datetime.strptime(s, fmt).date()
            except ValueError: pass
    return None


def _int(v):
    if v is None or v == "": return 0
    try: return int(float(v))
    except (TypeError, ValueError): return 0


def _str(v):
    return ("" if v is None else str(v)).strip()


# Workbook pivots include per-origin / per-date "TOTAL" rows that aggregate
# across the real warehouse ports. Treating these as a port double-counts the
# grand total. Filter them at every parser entry that consumes a port_code.
_AGGREGATE_PORT_NAMES = {"TOTAL", "GRAND TOTAL", "GRANDTOTAL", "ALL", "SUM", "TOT"}

def _is_aggregate_port(port_code: str) -> bool:
    return (port_code or "").strip().upper() in _AGGREGATE_PORT_NAMES


# ── Per-sheet parsers ─────────────────────────────────────────────────────────

def _bucket_by_date(sheet, *, date_col=0, skip_header_rows=1):
    """Yield (row_date, row_tuple) for each data row whose date is parseable."""
    rows_iter = sheet.iter_rows(values_only=True)
    for _ in range(skip_header_rows):
        next(rows_iter, None)
    for r in rows_iter:
        if not r:
            continue
        d = _to_date(r[date_col])
        if not d: continue
        yield d, r


def parse_sheet_7_ny(sheet, since: date):
    """Arabica certified — flat: date | origin | port_code | port_name | value | day | wow | mom | yoy.
    Drops zero-bag rows so the resulting JSON's by_port dicts stay sparse."""
    per_date: dict[date, dict[str, dict[str, int]]] = defaultdict(lambda: defaultdict(dict))
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        origin    = _str(r[1])
        port_code = _str(r[2])
        bags      = _int(r[4])
        if not origin or not port_code or bags <= 0: continue
        if _is_aggregate_port(port_code): continue
        per_date[d].setdefault(origin, {})[port_code] = bags
        n += 1
    print(f"    sheet 7_ny: {n} non-zero rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_5_pend(sheet, since: date):
    """Arabica pending grading — flat: date | port_code | port_name | bags | origin.
    Drops zero-bag rows."""
    per_date: dict[date, dict[str, dict[str, int]]] = defaultdict(lambda: defaultdict(dict))
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        port_code = _str(r[1])
        bags      = _int(r[3])
        origin    = _str(r[4])
        if not origin or not port_code or bags <= 0: continue
        if _is_aggregate_port(port_code): continue
        per_date[d].setdefault(origin, {})[port_code] = bags
        n += 1
    print(f"    sheet 5_ny_pend_gradings: {n} non-zero rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_6_gradings(sheet, since: date):
    """Arabica gradings — flat: date | port_code | port_name | status | bags | origin.
    Returns {date: {"passed": int, "failed": int,
                    "by_port_status": {port: {status: bags}},
                    "by_origin_status": {origin: {status: bags}},
                    "by_port_origin_status": {port: {origin: {status: bags}}}}}.
    The latter two are extracted from the workbook's `origin` column (was
    dropped pre-audit) so a future Arabica Graded drill can render
    port × origin × pass/fail rather than just port × pass/fail."""
    def _new_day():
        return {"passed": 0, "failed": 0,
                "by_port_status":        defaultdict(lambda: defaultdict(int)),
                "by_origin_status":      defaultdict(lambda: defaultdict(int)),
                "by_port_origin_status": defaultdict(lambda: defaultdict(lambda: defaultdict(int)))}
    per_date: dict[date, dict] = defaultdict(_new_day)
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        port_code = _str(r[1])
        status    = _str(r[3]).lower()
        bags      = _int(r[4])
        origin    = _str(r[5])
        if bags <= 0: continue
        if _is_aggregate_port(port_code): continue
        kind = "passed" if status.startswith("pass") else "failed" if status.startswith("fail") else None
        if kind is None: continue
        per_date[d][kind] += bags
        per_date[d]["by_port_status"][port_code][kind] += bags
        if origin:
            per_date[d]["by_origin_status"][origin][kind] += bags
            per_date[d]["by_port_origin_status"][port_code][origin][kind] += bags
        n += 1
    print(f"    sheet 6_ny_gradings: {n} rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_8a_issuers(sheet, since: date):
    """Arabica issuers — flat: date | port_code | port_name | issuers | value | type (MONTH/TODAY)."""
    per_date: dict[date, list[dict]] = defaultdict(list)
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        rec = {"port": _str(r[1]), "issuer": _str(r[3]), "value": _int(r[4]), "type": _str(r[5])}
        if not rec["issuer"] or rec["value"] == 0:  # drop zero rows
            if rec["value"] == 0: continue
            continue
        if _is_aggregate_port(rec["port"]): continue
        per_date[d].append(rec)
        n += 1
    print(f"    sheet 8a_ny_issuers: {n} non-zero rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_8b_stoppers(sheet, since: date):
    """Arabica receivers (stoppers) — flat: date | port_code | port_name | stoppers | value | type."""
    per_date: dict[date, list[dict]] = defaultdict(list)
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        rec = {"port": _str(r[1]), "stopper": _str(r[3]), "value": _int(r[4]), "type": _str(r[5])}
        if not rec["stopper"] or rec["value"] == 0: continue
        if _is_aggregate_port(rec["port"]): continue
        per_date[d].append(rec)
        n += 1
    print(f"    sheet 8b_ny_stoppers: {n} non-zero rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_12_age(sheet, since: date):
    """Arabica age detail — flat: month | date_actual | port_code | port_name | groups_of_day | as_of | bags | …
    Returns {date_actual: {port: [{age_bucket, bags}]}} — by date, by port, by age bucket."""
    per_date: dict[date, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    n = 0
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r: continue
        d = _to_date(r[1])
        if not d or d < since: continue
        port    = _str(r[2])
        bucket  = _str(r[4])
        bags    = _int(r[6])
        if not port or not bucket or bags <= 0: continue
        if _is_aggregate_port(port): continue
        per_date[d][port].append({"age_bucket": bucket, "bags": bags})
        n += 1
    print(f"    sheet 12_ny_age: {n} non-zero rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_9_ld(sheet, since: date):
    """Robusta stock report — flat: date | port_code | port_name | lotswithvalcert | lotsnontend | suspended | …
    Skips rows where all three lot counts are 0 (most ports on most days)."""
    per_date: dict[date, dict] = {}
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        port_code  = _str(r[1])
        port_name  = _str(r[2]) or None
        val_cert   = _int(r[3])
        non_tend   = _int(r[4])
        suspended  = _int(r[5])
        if not port_code: continue
        if _is_aggregate_port(port_code): continue
        if val_cert == 0 and non_tend == 0 and suspended == 0: continue
        ent = per_date.setdefault(d, {"cut_off_date": d.isoformat(), "ports": [],
                                       "grand_total": {"with_val_cert": 0, "non_tend": 0, "suspended": 0}})
        ent["ports"].append({"port_id": port_code, "port_name": port_name,
                              "with_val_cert": val_cert, "non_tend": non_tend, "suspended": suspended})
        ent["grand_total"]["with_val_cert"] += val_cert
        ent["grand_total"]["non_tend"]      += non_tend
        ent["grand_total"]["suspended"]     += suspended
        n += 1
    for ent in per_date.values():
        ent["ports"].sort(key=lambda p: -p["with_val_cert"])
    print(f"    sheet 9_ld: {n} rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_2_gradings(sheet, since: date):
    """Robusta gradings — flat: date | port_code | port_name | uk_cont_us | origin | class_num | tenderable | lot."""
    per_date: dict[date, dict] = defaultdict(lambda: {"entries": [],
                                                       "summary": {"tenderable_today": 0, "non_tenderable_today": 0,
                                                                   "lots_graded_today": 0,
                                                                   "tenderable_month": 0, "non_tenderable_month": 0,
                                                                   "lots_graded_month": 0}})
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        entry = {
            "port":             _str(r[1]),
            "section":          _str(r[3]),  # uk_cont_us
            "origin":           _str(r[4]).strip(),
            "class":            _int(r[5]) or None,
            "tenderable":       _str(r[6]).upper() == "Y",
            "lots":             _int(r[7]),
        }
        if entry["lots"] <= 0: continue
        if _is_aggregate_port(entry["port"]): continue
        per_date[d]["entries"].append(entry)
        if entry["tenderable"]:
            per_date[d]["summary"]["tenderable_today"] += entry["lots"]
        else:
            per_date[d]["summary"]["non_tenderable_today"] += entry["lots"]
        per_date[d]["summary"]["lots_graded_today"] += entry["lots"]
        n += 1
    print(f"    sheet 2_ld_gradings: {n} rows kept ({len(per_date)} dates).")
    return per_date


def parse_sheet_3_age(sheet, since_month_end: date):
    """Robusta age allowance — flat: date_actual | port_code | port_name | months_since_graded | value_mt | month_actual.
    Skips 0-MT cells so buckets[].by_port dicts stay sparse."""
    per_me: dict[date, dict] = defaultdict(lambda: {"valid": {"ports": set(), "buckets_raw": defaultdict(lambda: defaultdict(int))}})
    n = 0
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r: continue
        me = _to_date(r[0])
        if not me or me < since_month_end: continue
        port    = _str(r[1])
        months  = _int(r[3])
        mt      = _int(r[4])
        if not port or months <= 0 or mt <= 0: continue
        if _is_aggregate_port(port): continue
        per_me[me]["valid"]["ports"].add(port)
        per_me[me]["valid"]["buckets_raw"][months][port] += mt
        n += 1
    out: dict[date, dict] = {}
    for me, ent in per_me.items():
        ports_sorted = sorted(ent["valid"]["ports"])
        buckets = []
        grand = 0
        totals_by_port: dict[str, int] = defaultdict(int)
        for months in sorted(ent["valid"]["buckets_raw"].keys()):
            by_port = dict(ent["valid"]["buckets_raw"][months])
            grand_b = sum(by_port.values())
            buckets.append({"months_since_graded": months, "by_port": by_port, "grand_total_mt": grand_b,
                             "age_allowance_usd_per_mt": None})
            grand += grand_b
            for p, v in by_port.items(): totals_by_port[p] += v
        out[me] = {
            "report_date":   me.isoformat(),
            "valid":         {"report_date": me.isoformat(), "ports": ports_sorted,
                              "buckets": buckets, "totals_by_port": dict(totals_by_port),
                              "grand_total_mt": grand},
            "transition":    None,
        }
    print(f"    sheet 3_age_allowance: {n} rows kept ({len(out)} month-ends).")
    return out


def parse_sheet_4_iss_recv(sheet, since: date):
    """Robusta iss/recv — flat: date | commodity | member | origin | lots_sold | lots_bought."""
    per_date: dict[date, dict] = defaultdict(lambda: {"members": defaultdict(lambda: {"rows": [], "total_sold": 0, "total_bought": 0}),
                                                       "grand_total": {"sold": 0, "bought": 0}})
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        member = _str(r[2])
        origin = _str(r[3])
        sold   = _int(r[4])
        bought = _int(r[5])
        if not member: continue
        m = per_date[d]["members"][member]
        if sold or bought:
            m["rows"].append({"origin": origin, "sold": sold, "bought": bought})
        m["total_sold"]   += sold
        m["total_bought"] += bought
        per_date[d]["grand_total"]["sold"]   += sold
        per_date[d]["grand_total"]["bought"] += bought
        n += 1
    out: dict[date, dict] = {}
    for d, ent in per_date.items():
        out[d] = {
            "report_date":   d.isoformat(),
            "delivery_date": d.isoformat(),
            "members":       [{"code": code, **mdata} for code, mdata in ent["members"].items()],
            "grand_total":   ent["grand_total"],
        }
    print(f"    sheet 4_ld_issuer_receiver: {n} rows kept ({len(out)} dates).")
    return out


def parse_sheet_1_tenders(sheet, since: date):
    """Robusta tenders — PIVOTED: r1 is header banner, r2 has origin column headers, then date rows."""
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)                                        # r1 — banner
    header = next(rows, None) or ()
    origins = [_str(c) for c in header[1:] if _str(c) and _str(c).lower() != "row labels"]
    # Column index of each origin (offset by +1 because col 0 is the date).
    col_origins = []
    for i in range(1, len(header)):
        name = _str(header[i])
        if name and name.lower() != "grand total":
            col_origins.append((i, name))

    per_date: dict[date, dict] = {}
    n = 0
    for r in rows:
        if not r: continue
        d = _to_date(r[0])
        if not d or d < since: continue
        by_origin = []
        total_today = 0
        for ci, name in col_origins:
            v = _int(r[ci]) if ci < len(r) else 0
            if v:
                by_origin.append({"origin": name, "originals_today": v, "retenders_today": 0})
                total_today += v
        per_date[d] = {
            "report_date":     d.isoformat(),
            "delivery_period": None,
            "business_date":   d.isoformat(),
            "by_origin":       by_origin,
            "totals_today":    {"originals": total_today, "retenders": 0},
            "grand_totals":    {"originals": total_today, "retenders": 0},
        }
        n += 1
    print(f"    sheet 1_ld_tender: {n} date rows ({len(origins)} origin columns).")
    return per_date


# ── Deep history (chunked by 5-year buckets, light shape) ───────────────────
# Long-horizon chart background data. Same source sheets, but we keep only
# headline + by-port totals (no by_origin / by_group hierarchy) so each
# 5-year file stays well under a MB.

def _five_year_bucket(d: date) -> tuple[int, int]:
    start = (d.year // 5) * 5
    return (start, start + 4)


def build_deep_arabica_chunks(sheet_7) -> dict[tuple[int, int], list[dict]]:
    """Walk the FULL arabica certified sheet, bucket by (date, port) → totals,
    chunk by 5-year window. Returns {(start_yr, end_yr): [light_snap, …]}."""
    per_date_pp: dict[date, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    n = 0
    for d, r in _bucket_by_date(sheet_7):
        port_code = _str(r[2])
        bags      = _int(r[4])
        if not port_code or bags <= 0: continue
        if _is_aggregate_port(port_code): continue
        per_date_pp[d][port_code] += bags
        n += 1
    chunks: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for d, port_bags in per_date_pp.items():
        chunks[_five_year_bucket(d)].append({
            "date":           d.isoformat(),
            "total_bags":     sum(port_bags.values()),
            "by_port_totals": dict(port_bags),
        })
    for snaps in chunks.values():
        snaps.sort(key=lambda s: s["date"])
    print(f"    deep arabica: {n} non-zero certified rows → {len(per_date_pp)} dates, "
          f"{len(chunks)} 5-year chunks.")
    return chunks


def build_deep_robusta_chunks(sheet_9) -> dict[tuple[int, int], list[dict]]:
    """Walk the FULL robusta stock_report sheet, build light snapshots."""
    per_date: dict[date, dict] = {}
    n = 0
    for d, r in _bucket_by_date(sheet_9):
        port_code = _str(r[1])
        val_cert  = _int(r[3])
        if not port_code: continue
        if _is_aggregate_port(port_code): continue
        ent = per_date.setdefault(d, {"by_port_lots": {}, "total_lots_certified": 0})
        if val_cert > 0:
            ent["by_port_lots"][port_code] = val_cert
        ent["total_lots_certified"] += val_cert
        n += 1
    chunks: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for d, ent in per_date.items():
        chunks[_five_year_bucket(d)].append({
            "date":                 d.isoformat(),
            "total_lots_certified": ent["total_lots_certified"],
            "by_port_lots":         ent["by_port_lots"],
        })
    for snaps in chunks.values():
        snaps.sort(key=lambda s: s["date"])
    print(f"    deep robusta: {n} rows → {len(per_date)} dates, "
          f"{len(chunks)} 5-year chunks.")
    return chunks


def write_deep_chunks(market: str, chunks: dict[tuple[int, int], list[dict]]) -> None:
    """Write one JSON per 5-year bucket. Filenames:
        certified_stocks_<market>_deep_<startYr>-<endYr>.json
    Existing files are overwritten (deep is regenerated from the workbook
    each import; the live scraper never touches these)."""
    now_iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
    for (s_yr, e_yr), snaps in sorted(chunks.items()):
        path = OUT_DIR / f"certified_stocks_{market}_deep_{s_yr}-{e_yr}.json"
        payload = {
            "generated_at": now_iso,
            "market":       market,
            "bucket_years": [s_yr, e_yr],
            "shape":        "light (date + total + by_port_totals; no by_origin/by_group)",
            "snapshots":    snaps,
        }
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        print(f"  wrote {path.name}  ({path.stat().st_size:,} B; {len(snaps)} snapshots)")


def parse_sheet_13_overview(sheet, since: date):
    """Robusta grading queue+forecast — flat: date_actual | type | description | value."""
    per_date: dict[date, dict] = defaultdict(lambda: {"queue_lots": 0, "forecast_lots": 0, "total_pending_lots": 0})
    n = 0
    for d, r in _bucket_by_date(sheet):
        if d < since: continue
        kind  = _str(r[1]).lower()
        value = _int(r[3])
        if "queue" in kind:
            per_date[d]["queue_lots"] = value
        elif "forecast" in kind or "garding forecast" in kind:
            per_date[d]["forecast_lots"] = value
        per_date[d]["total_pending_lots"] = per_date[d]["queue_lots"] + per_date[d]["forecast_lots"]
        n += 1
    out = {d: {"report_date": d.isoformat(), **v} for d, v in per_date.items()}
    print(f"    sheet 13_ld_gradings_queue: {n} rows ({len(out)} dates).")
    return out


# ── Snapshot composition ──────────────────────────────────────────────────────

def _build_arabica_section(per_date_oo_pp: dict, d: date) -> dict | None:
    """origin → port → bags for one date → {by_port, by_group, by_origin, grand_total}."""
    block = per_date_oo_pp.get(d)
    if not block: return None
    by_origin: dict[str, dict] = {}
    by_port: dict[str, int] = defaultdict(int)
    by_group: dict[str, int] = defaultdict(int)
    grand = 0
    for origin, port_bags in block.items():
        total = sum(port_bags.values())
        if total == 0: continue
        grp = group_of(origin) or "Unknown"
        by_origin[origin] = {"by_port": dict(port_bags), "group": grp, "total": total}
        for p, b in port_bags.items(): by_port[p] += b
        by_group[grp] += total
        grand += total
    return {"grand_total": grand, "by_port": dict(by_port), "by_group": dict(by_group), "by_origin": by_origin}


def build_arabica_snapshots(certified, pending, gradings_today,
                             issuers_per_date, stoppers_per_date, age_per_date):
    dates = sorted(set(certified) | set(pending))
    snaps: list[dict] = []
    for d in dates:
        tc = _build_arabica_section(certified, d) or {"grand_total": 0, "by_port": {}, "by_group": {}, "by_origin": {}}
        pg = _build_arabica_section(pending, d)   or {"grand_total": 0, "by_port": {}, "by_group": {}, "by_origin": {}}
        gt = gradings_today.get(d, {})
        passed = gt.get("passed", 0); failed = gt.get("failed", 0)
        # Per-snapshot issuers/stoppers (TODAY type only — that's the daily flow;
        # MONTH type is cumulative-to-date and stored in latest_detail).
        iss_today = [{"port": e["port"], "issuer": e["issuer"], "value": e["value"]}
                     for e in issuers_per_date.get(d, [])
                     if e.get("type") == "TODAY" and e.get("value", 0) > 0]
        stp_today = [{"port": e["port"], "stopper": e["stopper"], "value": e["value"]}
                     for e in stoppers_per_date.get(d, [])
                     if e.get("type") == "TODAY" and e.get("value", 0) > 0]
        # Per-snapshot gradings detail — by port/origin/pass-fail. Captured
        # exhaustively from sheet 6 so future drills (Arabica Graded → origin)
        # work without re-importing.
        def _materialise(dd):
            """Convert nested defaultdicts into plain dicts (JSON-safe)."""
            if isinstance(dd, dict):
                return {k: _materialise(v) for k, v in dd.items()}
            return dd
        grad_origin   = _materialise(gt.get("by_origin_status", {}))
        grad_port_org = _materialise(gt.get("by_port_origin_status", {}))
        snaps.append({
            "date":                 d.isoformat(),
            "report_date":          d.isoformat(),
            "total_bags":           tc["grand_total"],
            "transition_bags":      0,
            "pending_grading_bags": pg["grand_total"],
            "rebagging_bags":       0,
            "passed_today_bags":    passed,
            "failed_today_bags":    failed,
            "by_port":              tc["by_port"],
            "by_group":             tc["by_group"],
            "sections": {
                "total_certified": tc,
                "pending_grading": pg,
            },
            # NEW per-snapshot fields — drive arabica Issued period view + drill.
            "issuers_today":        iss_today,
            "stoppers_today":       stp_today,
            "issued_total_today":   sum(e["value"] for e in iss_today),
            "received_total_today": sum(e["value"] for e in stp_today),
            # Gradings-today detail (origin-resolved).
            "graded_today_by_origin":      grad_origin,
            "graded_today_by_port_origin": grad_port_org,
        })
    latest_date = dates[-1] if dates else None
    latest_detail = None
    if latest_date:
        # age_detail is published at month-ends in sheet 12, not daily —
        # so latest_detail.age_detail uses the most recent age entry we have,
        # which may pre-date the snapshot's latest_date.
        latest_age_date = max(age_per_date) if age_per_date else None
        latest_detail = {
            "report_date":    latest_date.isoformat(),
            "total_certified": _build_arabica_section(certified, latest_date),
            "transition":      None,
            "pending_grading": _build_arabica_section(pending, latest_date),
            "rebagging":       None,
            "grading_today":  {"passed_today_bags": gradings_today.get(latest_date, {}).get("passed", 0),
                                "failed_today_bags": gradings_today.get(latest_date, {}).get("failed", 0),
                                "passed_text":       None, "failed_text": None},
            # Additive: NEW arabica-specific fields (don't break old readers).
            "issuers":         issuers_per_date.get(latest_date, []),
            "stoppers":        stoppers_per_date.get(latest_date, []),
            "age_detail":      age_per_date.get(latest_age_date) if latest_age_date else {},
            "age_detail_date": latest_age_date.isoformat() if latest_age_date else None,
        }
    return snaps, latest_detail, latest_date


def build_robusta_snapshots(stock, gradings, iss_recv, tenders, overview):
    # Only build snapshots for dates that have stock data. Other event dates
    # (gradings/iss-recv/tenders on days without a stock report) still appear
    # in recent_activity; we just don't synthesize a fake 0-stock snapshot
    # which made the time-series chart drop to 0 ~30 times across the year.
    all_dates = sorted(stock.keys())
    snaps: list[dict] = []
    for d in all_dates:
        st = stock.get(d) or {}
        gt = gradings.get(d) or {"summary": {"lots_graded_today": 0, "tenderable_today": 0, "non_tenderable_today": 0}}
        ir = iss_recv.get(d) or {"grand_total": {"sold": 0, "bought": 0}}
        tn = tenders.get(d) or {"totals_today": {"originals": 0, "retenders": 0}}
        snaps.append({
            "date":                 d.isoformat(),
            "cut_off_date":         st.get("cut_off_date", d.isoformat()),
            "total_lots_certified": (st.get("grand_total") or {}).get("with_val_cert", 0),
            "non_tend_lots":        (st.get("grand_total") or {}).get("non_tend", 0),
            "suspended_lots":       (st.get("grand_total") or {}).get("suspended", 0),
            "lots_graded_today":    gt["summary"]["lots_graded_today"],
            "lots_sold_today":      (ir.get("grand_total") or {}).get("sold", 0),
            "lots_bought_today":    (ir.get("grand_total") or {}).get("bought", 0),
            "tenders_today":        (tn.get("totals_today") or {}).get("originals", 0),
            "by_port_lots":         {p["port_id"]: p["with_val_cert"] for p in (st.get("ports") or [])},
        })
    return snaps, all_dates[-1] if all_dates else None


# ── Main run ──────────────────────────────────────────────────────────────────

def run(xlsx_path: Path, daily_keep_days: int, monthly_keep_months: int, write: bool):
    today = date.today()
    since_daily = today - timedelta(days=daily_keep_days)
    since_month_end = date(today.year, today.month, 1) - timedelta(days=monthly_keep_months * 31)

    print(f"=== loading workbook: {xlsx_path} ===")
    print(f"    daily window:   >= {since_daily.isoformat()}  ({daily_keep_days} days)")
    print(f"    monthly window: >= {since_month_end.isoformat()}  ({monthly_keep_months} months)\n")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    print("=== parsing arabica sheets ===")
    a_certified = parse_sheet_7_ny(wb["7_ny"], since_daily)
    a_pending   = parse_sheet_5_pend(wb["5_ny_pend_gradings"], since_daily)
    a_gradings  = parse_sheet_6_gradings(wb["6_ny_gradings"], since_daily)
    a_issuers   = parse_sheet_8a_issuers(wb["8a_ny_issuers"], since_daily)
    a_stoppers  = parse_sheet_8b_stoppers(wb["8b_ny_stoppers"], since_daily)
    a_age       = parse_sheet_12_age(wb["12_ny_age"], since_daily)

    print("\n=== parsing robusta sheets ===")
    r_stock      = parse_sheet_9_ld(wb["9_ld"], since_daily)
    r_gradings   = parse_sheet_2_gradings(wb["2_ld_gradings"], since_daily)
    r_iss_recv   = parse_sheet_4_iss_recv(wb["4_ld_issuer_receiver"], since_daily)
    r_tenders    = parse_sheet_1_tenders(wb["1_ld_tender"], since_daily)
    r_overview   = parse_sheet_13_overview(wb["13_ld_gradings_queue"], since_daily)
    r_age_allow  = parse_sheet_3_age(wb["3_age_allowance"], since_month_end)

    print("\n=== composing arabica JSON ===")
    a_snaps, a_latest_detail, a_latest_date = build_arabica_snapshots(
        a_certified, a_pending, a_gradings, a_issuers, a_stoppers, a_age,
    )
    print(f"    arabica snapshots: {len(a_snaps)}")

    print("\n=== composing robusta JSON ===")
    r_snaps, r_latest_date = build_robusta_snapshots(r_stock, r_gradings, r_iss_recv, r_tenders, r_overview)
    print(f"    robusta snapshots: {len(r_snaps)}")

    # Activity feeds (flatten the dict-by-date structures into the JSON shape).
    r_gradings_list   = [{"date": d.isoformat(), **v} for d, v in sorted(r_gradings.items())]
    r_iss_recv_list   = [{"date": d.isoformat(), **v} for d, v in sorted(r_iss_recv.items())]
    r_tenders_list    = [{"date": d.isoformat(), **v} for d, v in sorted(r_tenders.items())]
    r_overview_list   = [{"date": d.isoformat(), **v} for d, v in sorted(r_overview.items())]
    r_age_allow_list  = [{"month_end": me.isoformat(), **v} for me, v in sorted(r_age_allow.items(), reverse=True)]

    now_iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
    arabica_json = {
        "generated_at":  now_iso,
        "as_of":         a_latest_date.isoformat() if a_latest_date else None,
        "source_url":    "imported from synthesis xlsx",
        "snapshots":     a_snaps,
        "latest_detail": a_latest_detail,
        "errors":        [],
    }
    robusta_json = {
        "generated_at":  now_iso,
        "as_of":         r_latest_date.isoformat() if r_latest_date else None,
        "snapshots":     r_snaps,
        "latest_detail": {"stock_report": r_stock.get(r_latest_date) if r_latest_date else None,
                          "stock_report_url": "imported from synthesis xlsx"},
        "recent_activity": {
            "gradings":          r_gradings_list,
            "grading_appeals":   [],          # not in this workbook
            "iss_recv_daily":    r_iss_recv_list,
            "tenders":           r_tenders_list,
            "grading_overview":  r_overview_list,
            "infested_warrants": [],          # not in this workbook
        },
        "monthly": {
            "iss_recv_monthly": [],            # not split in this workbook (daily has it)
            "age_allowance":    r_age_allow_list,
        },
    }

    a_path = OUT_DIR / "certified_stocks_arabica.json"
    r_path = OUT_DIR / "certified_stocks_robusta.json"
    if a_path.exists():
        try:
            existing = json.loads(a_path.read_text(encoding="utf-8"))
            n_old = len(existing.get("snapshots") or [])
            arabica_json = _merge_arabica(arabica_json, existing)
            print(f"\n[merge] arabica: {n_old} existing snapshots → {len(arabica_json['snapshots'])} after merge")
        except Exception as e:  # noqa: BLE001
            print(f"[merge] arabica skipped ({e})")
    if r_path.exists():
        try:
            existing = json.loads(r_path.read_text(encoding="utf-8"))
            n_old = len(existing.get("snapshots") or [])
            robusta_json = _merge_robusta(robusta_json, existing)
            print(f"[merge] robusta: {n_old} existing snapshots → {len(robusta_json['snapshots'])} after merge")
        except Exception as e:  # noqa: BLE001
            print(f"[merge] robusta skipped ({e})")

    # Post-merge: drop synthetic zero-stock robusta snapshots that older runs
    # left behind (a date with no stock_report.csv but with iss-recv/grading
    # events used to land as total=0 — creating false dips in the chart).
    before = len(robusta_json.get("snapshots") or [])
    robusta_json["snapshots"] = [s for s in (robusta_json.get("snapshots") or [])
                                  if (s.get("total_lots_certified") or 0) > 0]
    dropped = before - len(robusta_json["snapshots"])
    if dropped:
        print(f"[filter] dropped {dropped} robusta snapshots with 0 stock (no source data those days)")

    if write:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        a_path.write_text(json.dumps(arabica_json, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        r_path.write_text(json.dumps(robusta_json, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        a_size = a_path.stat().st_size; r_size = r_path.stat().st_size
        print(f"\n=== wrote {a_path}  ({a_size:,} bytes)")
        print(f"=== wrote {r_path}  ({r_size:,} bytes)")

        # ── Deep history chunks (full workbook range, 5-year buckets, light shape) ──
        print("\n=== building deep history chunks (5-year buckets, light shape) ===")
        a_chunks = build_deep_arabica_chunks(wb["7_ny"])
        r_chunks = build_deep_robusta_chunks(wb["9_ld"])
        print(f"\n=== writing deep arabica chunks ===")
        write_deep_chunks("arabica", a_chunks)
        print(f"=== writing deep robusta chunks ===")
        write_deep_chunks("robusta", r_chunks)

    print(f"\nSUMMARY:")
    print(f"  arabica: {len(arabica_json['snapshots'])} snapshots · as_of={arabica_json['as_of']}")
    print(f"  robusta: {len(robusta_json['snapshots'])} snapshots · as_of={robusta_json['as_of']}")
    ra = robusta_json["recent_activity"]
    print(f"  robusta events: gradings={len(ra['gradings'])} iss/recv={len(ra['iss_recv_daily'])} "
          f"tenders={len(ra['tenders'])} overview={len(ra['grading_overview'])}")
    print(f"  robusta age-allowance months: {len(robusta_json['monthly']['age_allowance'])}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--days",   type=int, default=DAILY_KEEP_DAYS,
                    help="last N days to keep in daily snapshots (default 365)")
    ap.add_argument("--months", type=int, default=MONTHLY_KEEP_MONTHS,
                    help="last N months to keep in monthly snapshots (default 24)")
    ap.add_argument("--no-write", action="store_true")
    args = ap.parse_args()
    if not args.xlsx.exists():
        print(f"ERROR: {args.xlsx} not found", file=sys.stderr); sys.exit(1)
    run(args.xlsx, args.days, args.months, write=not args.no_write)


if __name__ == "__main__":
    main()
