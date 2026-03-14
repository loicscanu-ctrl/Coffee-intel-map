#!/usr/bin/env python3
"""
One-time import: reads NY COT + LDN COT + Other sheets from the Excel file
and upserts all rows into the cot_weekly table.

Usage:
    cd backend
    DATABASE_URL=postgresql://... python import_cot_excel.py \
        --file "C:/Users/Loic Scanu/OneDrive - Tuan Loc Commodities/TradeTeam/COT report.xlsx"
"""
import argparse
import bisect
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))


def _to_float(val):
    """Coerce openpyxl cell value to float or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_int(val):
    """Coerce openpyxl cell value to int or None."""
    f = _to_float(val)
    return int(f) if f is not None else None


def _build_prices(ws_other):
    """
    Read the Other sheet (data starts at row 3, openpyxl min_row=3).
    Returns (sorted_dates, prices_dict) where prices_dict maps date -> field dict.

    Column mapping (0-indexed from row tuple):
      col 0: Date          col 1: price_ny     col 4: price_ldn
      col 7: structure_ny  col 8: structure_ldn
      col 18: exch_oi_ny   col 19: exch_oi_ldn
      col 20: vol_ny       col 21: vol_ldn
      col 22: efp_ny       col 23: efp_ldn
      col 24: spread_vol_ny col 25: spread_vol_ldn
    """
    prices = {}
    for row in ws_other.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        try:
            d = row[0].date() if hasattr(row[0], "date") else None
        except Exception:
            continue
        if d is None:
            continue
        prices[d] = {
            "price_ny":       _to_float(row[1]),
            "price_ldn":      _to_float(row[4]),
            "structure_ny":   _to_float(row[7]),
            "structure_ldn":  _to_float(row[8]),
            "exch_oi_ny":     _to_int(row[18]),
            "exch_oi_ldn":    _to_int(row[19]),
            "vol_ny":         _to_int(row[20]),
            "vol_ldn":        _to_int(row[21]),
            "efp_ny":         _to_float(row[22]),
            "efp_ldn":        _to_float(row[23]),
            "spread_vol_ny":  _to_float(row[24]),
            "spread_vol_ldn": _to_float(row[25]),
        }
    sorted_dates = sorted(prices.keys())
    return sorted_dates, prices


def _lookup_price(report_date, sorted_dates, prices):
    """Return price fields for the most recent date <= report_date."""
    if not sorted_dates:
        return {}
    idx = bisect.bisect_right(sorted_dates, report_date) - 1
    if idx < 0:
        return {}
    return prices[sorted_dates[idx]]


def _import_sheet(ws, market: str, ny_col_map: bool, sorted_dates, prices):
    """
    Import one COT sheet (NY or LDN). Returns count of rows processed.

    Column mapping (0-indexed):
      NY COT:                         LDN COT:
        0  date                         0  date
        1  oi_total                     1  oi_total
        2  pmpu_long  ...              (same cols 2-14)
       32  t_pmpu_long                 17  t_pmpu_long
       33  t_pmpu_short                18  t_pmpu_short
       34  t_swap_long                 19  t_swap_long
       35  t_swap_short                20  t_swap_short
       36  t_swap_spread               21  t_swap_spread
       37  t_mm_long                   22  t_mm_long
       38  t_mm_short                  23  t_mm_short
       39  t_mm_spread                 24  t_mm_spread
       40  t_other_long                25  t_other_long
       41  t_other_short               26  t_other_short
       42  t_other_spread              27  t_other_spread
       43  t_nr_long                   N/A (None)
       44  t_nr_short                  N/A (None)
    """
    from scraper.db import upsert_cot_weekly

    tc = 32 if ny_col_map else 17
    has_nr_traders = ny_col_map

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            d = row[0].date() if hasattr(row[0], "date") else None
        except Exception:
            continue
        if d is None:
            continue

        fields = {
            "oi_total":     _to_int(row[1]),
            "pmpu_long":    _to_int(row[2]),
            "pmpu_short":   _to_int(row[3]),
            "swap_long":    _to_int(row[4]),
            "swap_short":   _to_int(row[5]),
            "swap_spread":  _to_int(row[6]),
            "mm_long":      _to_int(row[7]),
            "mm_short":     _to_int(row[8]),
            "mm_spread":    _to_int(row[9]),
            "other_long":   _to_int(row[10]),
            "other_short":  _to_int(row[11]),
            "other_spread": _to_int(row[12]),
            "nr_long":      _to_int(row[13]),
            "nr_short":     _to_int(row[14]),
            "t_pmpu_long":    _to_int(row[tc + 0]),
            "t_pmpu_short":   _to_int(row[tc + 1]),
            "t_swap_long":    _to_int(row[tc + 2]),
            "t_swap_short":   _to_int(row[tc + 3]),
            "t_swap_spread":  _to_int(row[tc + 4]),
            "t_mm_long":      _to_int(row[tc + 5]),
            "t_mm_short":     _to_int(row[tc + 6]),
            "t_mm_spread":    _to_int(row[tc + 7]),
            "t_other_long":   _to_int(row[tc + 8]),
            "t_other_short":  _to_int(row[tc + 9]),
            "t_other_spread": _to_int(row[tc + 10]),
            "t_nr_long":  _to_int(row[tc + 11]) if has_nr_traders else None,
            "t_nr_short": _to_int(row[tc + 12]) if has_nr_traders else None,
        }

        fields.update(_lookup_price(d, sorted_dates, prices))
        upsert_cot_weekly(market, d, fields)
        count += 1

    return count


def main():
    parser = argparse.ArgumentParser(description="Import CoT Excel data into cot_weekly table")
    parser.add_argument("--file", required=True, help="Path to COT report.xlsx")
    args = parser.parse_args()

    import openpyxl
    print(f"Opening {args.file} ...")
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)

    print("Reading Other sheet for prices/volumes ...")
    ws_other = wb["Other"]
    sorted_dates, prices = _build_prices(ws_other)
    print(f"  Other sheet: {len(prices)} date entries loaded")

    print("Importing NY COT ...")
    ws_ny = wb["NY COT"]
    ny_count = _import_sheet(ws_ny, "ny", ny_col_map=True,
                              sorted_dates=sorted_dates, prices=prices)
    print(f"  NY: {ny_count} rows upserted")

    print("Importing LDN COT ...")
    ws_ldn = wb["LDN COT"]
    ldn_count = _import_sheet(ws_ldn, "ldn", ny_col_map=False,
                               sorted_dates=sorted_dates, prices=prices)
    print(f"  LDN: {ldn_count} rows upserted")

    print(f"Done. Total: {ny_count + ldn_count} rows.")


if __name__ == "__main__":
    main()
