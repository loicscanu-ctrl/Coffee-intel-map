"""
CONAB monthly scraper.
Fetches Brazil coffee production data from two sources:
  1. Safra bulletin — acreage + yield KPIs (gov.br, XLS via xlrd)
  2. Custos de Produção report — production cost breakdown (gov.br, Playwright for JS tabs)

URLs migrated to gov.br in 2024.
"""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import date, datetime

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Listing page — contains links to individual levantamentos
CONAB_SAFRA_LISTING_URL = (
    "https://www.gov.br/conab/pt-br/atuacao/informacoes-agropecuarias/safras/safra-de-cafe"
)
# Planilhas page — uses JS tabs (Agrícolas tab → coffee xlsx links)
CONAB_CUSTOS_URL = (
    "https://www.gov.br/conab/pt-br/atuacao/informacoes-agropecuarias/"
    "custos-de-producao/planilhas-de-custos-de-producao"
)

_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; CoffeeIntelScraper/1.0)"}

# Cost component aggregation: regex pattern → (English label, hex color)
# Each matching row is summed into its category bucket.
# Labels come from CONAB historical series (e.g. "6 - Mão de obra", "9 - Fertilizantes")
_COST_COMPONENT_MAP = {
    r"fertilizante|agrot[oó]xico|semente|insumo": ("Inputs",        "#3b82f6"),
    r"m[ãa]o.de.obra|m\.o\.":                    ("Labor",         "#22c55e"),
    r"m[aá]quina|trator|colheita|mecaniza":       ("Mechanization", "#f59e0b"),
    r"arrendamento":                              ("Land rent",     "#8b5cf6"),
    r"administra":                                ("Admin",         "#475569"),
}

# Input detail mapping: regex → English label
_INPUT_DETAIL_MAP = {
    r"nitrog|ur[eé]ia|sal.amonio": "Nitrogen (urea / AN)",
    r"pot[áa]ssio|kcl":           "Potassium (KCl)",
    r"f[oó]sforo|\bmap\b|\bdap\b": "Phosphorus (MAP)",
    r"defensivo|pesticida":        "Pesticides / fungicides",
    r"cal[cç][aá]r|lime|corretivo": "Lime / soil correction",
}


# ---------------------------------------------------------------------------
# Pure parsing helpers (no I/O, fully testable)
# ---------------------------------------------------------------------------

def _brl_to_float(s: str) -> float | None:
    """Convert Brazilian number format '2.240,5' → 2240.5. Return None on empty/invalid."""
    s = s.strip()
    if not s:
        return None
    try:
        cleaned = s.replace(".", "").replace(",", ".")
        return float(cleaned)
    except (ValueError, AttributeError):
        return None


def _current_season() -> str:
    """Return current crop-year label, e.g. '2025/26'. Brazil crop year starts July."""
    today = date.today()
    if today.month >= 7:
        return f"{today.year}/{str(today.year + 1)[-2:]}"
    else:
        return f"{today.year - 1}/{str(today.year)[-2:]}"


def _parse_conab_safra_xls(content: bytes) -> dict | None:
    """Parse CONAB Safra XLS (old .xls format) using xlrd.

    Expects sheet '1 Café Total' with columns:
      0: REGIÃO/UF name
      1: Área Safra 2025 (ha) — previous
      2: Área Safra 2026 (ha) — current estimate
      4: Produtividade Safra 2025 (sc/ha)
      5: Produtividade Safra 2026 (sc/ha)

    Looks for row where col 0 (uppercased) == 'BRASIL'.
    Returns {"harvested_area_kha": float, "yield_bags_ha": float} or None.
    """
    import xlrd  # noqa: PLC0415

    try:
        wb = xlrd.open_workbook(file_contents=content)
    except Exception as e:
        logger.warning(f"[conab_supply] xlrd cannot open file: {e}")
        return None

    # Find the total sheet
    sheet_name = None
    for name in wb.sheet_names():
        if "total" in name.lower() and "caf" in name.lower() and "rea" not in name.lower():
            sheet_name = name
            break
    if sheet_name is None and wb.nsheets > 0:
        sheet_name = wb.sheet_names()[1]  # fallback to second sheet

    if sheet_name is None:
        logger.warning("[conab_supply] Safra XLS: no suitable sheet found")
        return None

    ws = wb.sheet_by_name(sheet_name)

    for row_idx in range(ws.nrows):
        first_cell = ws.cell_value(row_idx, 0)
        if not isinstance(first_cell, str):
            continue
        if first_cell.strip().upper() == "BRASIL":
            if ws.ncols < 6:
                continue
            # Col 2 = current-year area (ha), col 5 = current-year yield (sc/ha)
            area_ha = ws.cell_value(row_idx, 2)
            yld_sc_ha = ws.cell_value(row_idx, 5)
            if isinstance(area_ha, (int, float)) and isinstance(yld_sc_ha, (int, float)):
                return {
                    "harvested_area_kha": round(float(area_ha) / 1000, 1),
                    "yield_bags_ha":      round(float(yld_sc_ha), 2),
                }

    logger.warning("[conab_supply] Safra XLS: BRASIL row not found")
    return None


def _select_custos_sheet_xls(wb) -> object:
    """Pick the most recent Sul de Minas municipality sheet from the arabica historical series XLS."""
    _SULDEMINAS_MUNI = ["guaxup", "tr\xeas pontas", "patroc", "s.s. para", "tres pontas"]
    sul_sheets = []
    for name in wb.sheet_names():
        name_l = name.lower()
        if any(m in name_l for m in _SULDEMINAS_MUNI):
            year_match = re.search(r"(\d{4})", name)
            if year_match:
                sul_sheets.append((int(year_match.group(1)), name))

    if sul_sheets:
        sul_sheets.sort(key=lambda t: t[0], reverse=True)
        return wb.sheet_by_name(sul_sheets[0][1])

    for i, name in enumerate(wb.sheet_names()):
        if "ndice" not in name.lower():
            return wb.sheet_by_index(i)
    return wb.sheet_by_index(0)


def _select_conilon_sheet_xls(wb) -> object:
    """Pick the most recent Espírito Santo municipality sheet from the conilon historical series XLS."""
    _ES_MUNI = ["linhares", "colatina", "s\xe3o mateus", "sao mateus", "barra de s", "nova ven\xe9cia"]
    es_sheets = []
    for name in wb.sheet_names():
        name_l = name.lower()
        if any(m in name_l for m in _ES_MUNI):
            year_match = re.search(r"(\d{4})", name)
            if year_match:
                es_sheets.append((int(year_match.group(1)), name))

    if es_sheets:
        es_sheets.sort(key=lambda t: t[0], reverse=True)
        return wb.sheet_by_name(es_sheets[0][1])

    for i, name in enumerate(wb.sheet_names()):
        if "ndice" not in name.lower():
            return wb.sheet_by_index(i)
    return wb.sheet_by_index(0)


def _iter_custos_rows_xls(ws) -> list[tuple[str, float | None]]:
    """Yield (label, per_bag_value) pairs from a historical series XLS cost sheet.

    Column layout:
      0: DISCRIMINAÇÃO (label)
      1: CUSTO POR HA
      2: CUSTO / SACO (60 kg)   ← this is what we want
    """
    rows = []
    for row_idx in range(ws.nrows):
        label_raw = ws.cell_value(row_idx, 0)
        if not isinstance(label_raw, str):
            continue
        label = label_raw.strip()
        if not label:
            continue
        # Use col 2 (per-bag value) when available
        value: float | None = None
        if ws.ncols > 2:
            cell = ws.cell_value(row_idx, 2)
            if isinstance(cell, (int, float)) and float(cell) > 0:
                value = float(cell)
        # Fallback: col 1 (per-ha)
        if value is None and ws.ncols > 1:
            cell = ws.cell_value(row_idx, 1)
            if isinstance(cell, (int, float)) and float(cell) > 0:
                value = float(cell)
        rows.append((label, value))
    return rows


def _iter_custos_rows_xlsx(ws) -> list[tuple[str, float | None]]:
    """Yield (label, value) pairs from an xlsx cost sheet (old format).

    Column 0 is label; first positive numeric value across remaining cols is the value.
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label_raw = row[0]
        if not isinstance(label_raw, str) or not label_raw.strip():
            continue
        value: float | None = None
        for cell in row[1:]:
            if isinstance(cell, (int, float)) and cell is not None and cell > 0:
                value = float(cell)
                break
            if isinstance(cell, str):
                parsed = _brl_to_float(cell)
                if parsed is not None and parsed > 0:
                    value = parsed
                    break
        rows.append((label_raw.strip(), value))
    return rows


def _parse_conab_custos_excel(
    content: bytes,
    brl_usd: float,
    coffee_type: str = "arabica",
) -> dict | None:
    """Parse CONAB Custos Excel (.xlsx or .xls). Supports historical series XLS format.

    coffee_type: "arabica" → Sul de Minas sheet; "conilon" → Espírito Santo sheet.
    Returns None if no components or no total found.
    """
    is_xlsx = content[:2] == b"PK"

    if is_xlsx:
        import openpyxl  # noqa: PLC0415
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        if coffee_type == "conilon":
            # Conilon xlsx: try "Linhares" or "Espírito Santo" sheet, else active
            ws = next(
                (wb[n] for n in wb.sheetnames
                 if any(k in n.lower() for k in ["linhares", "colatina", "esp\xedrito", "conilon"])),
                wb.active,
            )
        else:
            ws = wb["Sul de Minas"] if "Sul de Minas" in wb.sheetnames else wb.active
        pairs = _iter_custos_rows_xlsx(ws)
        wb.close()
    else:
        import xlrd  # noqa: PLC0415
        wb = xlrd.open_workbook(file_contents=content)
        ws = _select_conilon_sheet_xls(wb) if coffee_type == "conilon" else _select_custos_sheet_xls(wb)
        pairs = _iter_custos_rows_xls(ws)

    # Aggregate into component buckets
    component_totals: dict[str, dict] = {}
    inputs_detail_brl: list[dict] = []
    total_brl: float | None = None

    for label_raw, value in pairs:
        if value is None:
            continue
        label_lower = label_raw.lower()

        if re.search(r"custo total|total geral", label_lower):
            total_brl = value
            continue

        matched = False
        for pattern, (eng_label, color) in _COST_COMPONENT_MAP.items():
            if re.search(pattern, label_lower):
                if eng_label not in component_totals:
                    component_totals[eng_label] = {"label": eng_label, "brl": 0.0, "color": color}
                component_totals[eng_label]["brl"] = round(
                    component_totals[eng_label]["brl"] + value, 2
                )
                matched = True
                break

        if matched:
            continue

        for pattern, eng_label in _INPUT_DETAIL_MAP.items():
            if re.search(pattern, label_lower):
                inputs_detail_brl.append({"label": eng_label, "brl": value})
                break

    components_brl = list(component_totals.values())

    if not components_brl or total_brl is None:
        return None

    return {
        "season":               _current_season(),
        "total_brl_per_bag":    total_brl,
        "prev_total_brl_per_bag": None,
        "brl_usd":              brl_usd,
        "components_brl":       components_brl,
        "inputs_detail_brl":    inputs_detail_brl,
        "source_label":         f"CONAB Custos {coffee_type.capitalize()} {date.today():%b %Y}",
    }


# ---------------------------------------------------------------------------
# DB-writing helpers
# ---------------------------------------------------------------------------

def _get_brl_usd() -> float:
    """Fetch current BRL/USD rate via yfinance ticker 'BRL=X'. Return 5.0 on any failure."""
    try:
        import yfinance as yf
        ticker = yf.Ticker("BRL=X")
        info = ticker.fast_info
        rate = float(info.last_price)
        if rate and rate > 0:
            return rate
        return 5.0
    except Exception:
        return 5.0


def _find_safra_xls_url() -> str | None:
    """Fetch CONAB Safra listing page, return direct download URL of the latest café XLS.

    Flow:
      1. Fetch listing page
      2. Find first link whose href contains 'levantamento-de-cafe'
      3. Strip last path component to get the levantamento folder URL
      4. Fetch that folder page; find the first link whose text ends with '.xls' or '.xlsx'
      5. Strip trailing '/view' from that href and return
    """
    try:
        resp = requests.get(CONAB_SAFRA_LISTING_URL, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        main = soup.find("main") or soup

        # Step 2: find first levantamento link
        lev_href = None
        for a in main.find_all("a", href=True):
            if "levantamento-de-cafe" in a["href"].lower():
                lev_href = a["href"]
                break

        if not lev_href:
            logger.warning("[conab_supply] Safra: no levantamento link found on listing page")
            return None

        # Step 3: strip last path component → folder URL
        folder_url = lev_href.rsplit("/", 1)[0]

        # Step 4: fetch folder page
        resp2 = requests.get(folder_url, headers=_HEADERS, timeout=30)
        resp2.raise_for_status()
        soup2 = BeautifulSoup(resp2.text, "html.parser")
        main2 = soup2.find("main") or soup2

        # Find XLS link (href or link text contains .xls)
        for a in main2.find_all("a", href=True):
            href = a["href"]
            text = a.get_text(strip=True)
            if ".xls" in href.lower() or ".xls" in text.lower():
                # Remove trailing /view if present
                download_url = re.sub(r"/view$", "", href)
                return download_url

        logger.warning(f"[conab_supply] Safra: no .xls link found in {folder_url}")
        return None

    except Exception as e:
        logger.warning(f"[conab_supply] _find_safra_xls_url error: {e}")
        return None


def _scrape_acreage_yield(db) -> None:
    """Fetch CONAB Safra XLS from gov.br, parse BRASIL row, upsert NewsItem."""
    import os
    import sys
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
    from sqlalchemy import delete

    from models import NewsItem

    try:
        xls_url = _find_safra_xls_url()
        if xls_url is None:
            logger.warning("[conab_supply] Safra: could not find XLS URL — retaining previous data")
            return

        resp = requests.get(xls_url, headers=_HEADERS, timeout=60)
        resp.raise_for_status()
        parsed = _parse_conab_safra_xls(resp.content)

        if parsed is None:
            logger.warning("[conab_supply] Safra: could not parse XLS — retaining previous data")
            return

        # Read previous values for YoY comparison
        prev_meta: dict = {}
        existing = db.query(NewsItem).filter(NewsItem.source == "CONAB Safra").first()
        if existing and existing.meta:
            try:
                prev_meta = json.loads(existing.meta)
            except (json.JSONDecodeError, TypeError):
                prev_meta = {}

        db.execute(delete(NewsItem).where(NewsItem.source == "CONAB Safra"))
        db.add(NewsItem(
            title=f"CONAB Acreage Yield – {_current_season()}",
            source="CONAB Safra",
            category="supply",
            tags=["conab", "supply", "brazil", "acreage"],
            meta=json.dumps({
                "season":              _current_season(),
                "harvested_area_kha":  parsed["harvested_area_kha"],
                "prev_area_kha":       prev_meta.get("harvested_area_kha"),
                "yield_bags_ha":       parsed["yield_bags_ha"],
                "prev_yield_bags_ha":  prev_meta.get("yield_bags_ha"),
                "source_label":        f"CONAB Safra {date.today():%b %Y}",
            }),
            pub_date=datetime.utcnow(),
        ))
        db.commit()
        print(f"[conab_supply] Safra OK — area={parsed['harvested_area_kha']} kha, "
              f"yield={parsed['yield_bags_ha']} bags/ha")

    except Exception as e:
        db.rollback()
        logger.error(f"[conab_supply] Safra FAILED: {e}")


def _find_custos_xls_url(coffee_type: str = "arabica") -> str | None:
    """Fetch CONAB Custos Agrícolas tab content (plain HTML, no JS needed).

    coffee_type: "arabica" or "conilon"
    Returns direct download URL for the matching historical series XLS.
    """
    agricolas_url = CONAB_CUSTOS_URL + "/copy_of_agricolas"
    try:
        resp = requests.get(agricolas_url, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        keywords = (
            ["conilon", "conillon"]
            if coffee_type == "conilon"
            else ["arabica", "ar\xe1bica", "ar\u00e1bica"]
        )

        for a in soup.find_all("a", href=True):
            href = a["href"]
            text = a.get_text(strip=True).lower()
            href_l = href.lower()
            if any(k in href_l or k in text for k in keywords):
                if ".xls" in href_l:
                    return re.sub(r"/view$", "", href)

        logger.warning(f"[conab_supply] Custos: no café {coffee_type} XLS link found")
        return None
    except Exception as e:
        logger.warning(f"[conab_supply] _find_custos_xls_url({coffee_type}) error: {e}")
        return None


async def _scrape_production_cost(page, db, coffee_type: str = "arabica") -> None:  # noqa: ARG001
    """Fetch CONAB Custos Agrícolas tab, find the matching coffee XLS, parse and store."""
    import os
    import sys
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
    from sqlalchemy import delete

    from models import NewsItem

    db_source = "CONAB Custos" if coffee_type == "arabica" else "CONAB Custos Conilon"

    try:
        brl_usd = _get_brl_usd()

        xlsx_url = _find_custos_xls_url(coffee_type)
        if xlsx_url is None:
            logger.warning(f"[conab_supply] Custos {coffee_type}: no XLS link — retaining previous data")
            return

        xls_resp = requests.get(xlsx_url, headers=_HEADERS, timeout=120)
        xls_resp.raise_for_status()

        prev_total: float | None = None
        existing = db.query(NewsItem).filter(NewsItem.source == db_source).first()
        if existing and existing.meta:
            try:
                prev_total = json.loads(existing.meta).get("total_brl_per_bag")
            except (json.JSONDecodeError, TypeError):
                prev_total = None

        parsed = _parse_conab_custos_excel(xls_resp.content, brl_usd, coffee_type=coffee_type)

        if parsed is None:
            logger.warning(f"[conab_supply] Custos {coffee_type}: could not parse Excel — retaining")
            return

        parsed["prev_total_brl_per_bag"] = prev_total

        db.execute(delete(NewsItem).where(NewsItem.source == db_source))
        db.add(NewsItem(
            title=f"CONAB Production Cost {coffee_type.capitalize()} – {_current_season()}",
            source=db_source,
            category="supply",
            tags=["conab", "supply", "brazil", "cost", coffee_type],
            meta=json.dumps(parsed),
            pub_date=datetime.utcnow(),
        ))
        db.commit()
        print(f"[conab_supply] Custos {coffee_type} OK — total={parsed['total_brl_per_bag']} BRL/bag")

    except Exception as e:
        db.rollback()
        logger.error(f"[conab_supply] Custos {coffee_type} FAILED: {e}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

async def run(page, db) -> None:
    """Monthly entry point. Called from run_monthly.py."""
    _scrape_acreage_yield(db)
    await _scrape_production_cost(page, db, coffee_type="arabica")
    await _scrape_production_cost(page, db, coffee_type="conilon")
