"""
Farmer Economics daily scraper.
Fetches: weather (Open-Meteo), ENSO/ONI (NOAA CPC), fertilizer prices (World Bank).
"""
from __future__ import annotations

import io
import json
from collections import defaultdict
from datetime import date, datetime

import requests

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REGIONS = [
    {"name": "Sul de Minas",   "lat": -21.76, "lon": -45.25},
    {"name": "Cerrado",        "lat": -18.50, "lon": -47.50},
    {"name": "Paraná",         "lat": -23.55, "lon": -51.17},
    {"name": "Espírito Santo", "lat": -20.08, "lon": -41.37},
]

OPEN_METEO_URL = (
    "https://api.open-meteo.com/v1/forecast"
    "?latitude={lat}&longitude={lon}"
    "&hourly=temperature_2m,dew_point_2m,cloud_cover,wind_speed_10m"
    ",precipitation,precipitation_probability,et0_fao_evapotranspiration"
    ",soil_moisture_0_to_1cm,soil_moisture_1_to_3cm,soil_moisture_3_to_9cm"
    ",soil_moisture_9_to_27cm,soil_moisture_27_to_81cm"
    "&forecast_days=14&timezone=America/Sao_Paulo"
)

# Soil constants for Brazilian Oxisol (Latossolo)
_PWP = 0.15   # Permanent Wilting Point (m³/m³)
_AWC = 0.20   # Available Water Capacity (m³/m³); field capacity ≈ 0.35

NOAA_ONI_URL = "https://www.cpc.ncep.noaa.gov/data/indices/oni.ascii.txt"

WORLD_BANK_COMMODITY_PAGE = "https://www.worldbank.org/en/research/commodity-markets"
WORLD_BANK_FALLBACK_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
)

# Season code → calendar month number
_SEASON_MONTH = {
    "DJF": 1, "JFM": 2, "FMA": 3, "MAM": 4,
    "AMJ": 5, "MJJ": 6, "JJA": 7, "JAS": 8,
    "ASO": 9, "SON": 10, "OND": 11, "NDJ": 12,
}

_MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

# ---------------------------------------------------------------------------
# Pure parsing helpers (no I/O, fully testable)
# ---------------------------------------------------------------------------

def _vpd_kpa(temp_c: float, dew_point_c: float) -> float:
    """Vapour pressure deficit in kPa, computed from temperature and dew point."""
    import math
    e_sat = 0.6108 * math.exp(17.27 * temp_c    / (temp_c    + 237.3))
    e_act = 0.6108 * math.exp(17.27 * dew_point_c / (dew_point_c + 237.3))
    return max(0.0, round(e_sat - e_act, 3))


def _frost_risk(min_temp: float, cloud_cover: float, wind_kmh: float, dew_point: float) -> str:
    """Estimate leaf-surface temperature and return frost risk (H/M/L/-).

    Radiation cooling lowers the surface below the 2m air temperature on
    clear, calm nights. Formula:
        T_surface = min_temp
                    − (1 − cloud/100) × max(0, 5 − 0.4 × wind_kmh)
                    − 0.5  if dew_point < 2°C  (dry air amplifies cooling)

    Thresholds (°C):  < 0 → H,  0–3 → M,  3–6 → L,  ≥ 6 → —
    """
    radiation_cooling = (1 - cloud_cover / 100) * max(0.0, 5.0 - 0.4 * wind_kmh)
    dew_correction    = 0.5 if dew_point < 2.0 else 0.0
    t_surface = min_temp - radiation_cooling - dew_correction
    if t_surface < 0:
        return "H"
    if t_surface < 3:
        return "M"
    if t_surface < 6:
        return "L"
    return "-"


def _root_zone_moisture(sm_0_9: float, sm_9_27: float, sm_27_81: float) -> float:
    """Weighted root-zone soil moisture for coffee (m³/m³).

    Uses Open-Meteo forecast depth bands (deepest available: 27–81 cm).
    Weights reflect arabica root distribution in Brazilian red latosol:
        0–9 cm   (10%): surface — dries quickly, low root density
        9–27 cm  (25%): shallow root zone
        27–81 cm (65%): engine room + deep reserve (combined, deepest layer available)
    """
    return sm_0_9 * 0.10 + sm_9_27 * 0.25 + sm_27_81 * 0.65


def _drought_score(precip_mm: float, et0_mm: float, root_zone_sm: float) -> float:
    """Compute a raw daily drought stress score (0–4+).

    Agronomic model validated for arabica/conilon in Brazilian Oxisol:

    Step 1 — RWC-based base score (soil is the true water bank):
        RWC = (SM − PWP) / AWC × 100      [0–100%]
        ≥ 80 → 0.0  (comfortable)
        ≥ 60 → 0.5
        ≥ 40 → 1.0
        ≥ 20 → 2.0
        <  20 → 3.0  (near wilting)

    Step 2 — ET0 atmospheric demand adjustment (Penman-Monteith):
        > 7 mm/day → +0.5   (extreme evaporative pull)
        > 5 mm/day → +0.25  (elevated demand)

    Step 3 — Precipitation relief:
        Effective rain = actual_mm × effectiveness (0.20 / 0.60 / 0.90 by intensity)
        relief = min(1.5, effective / 5)

    Raw score = max(0, base + et0_adj − relief)
    """
    # Step 1 — RWC
    rwc = max(0.0, (root_zone_sm - _PWP) / _AWC * 100.0)
    if rwc >= 80:
        base = 0.0
    elif rwc >= 60:
        base = 0.5
    elif rwc >= 40:
        base = 1.0
    elif rwc >= 20:
        base = 2.0
    else:
        base = 3.0

    # Step 2 — ET0 driver
    if et0_mm > 7.0:
        et0_adj = 0.5
    elif et0_mm > 5.0:
        et0_adj = 0.25
    else:
        et0_adj = 0.0

    # Step 3 — precipitation relief
    if precip_mm < 3.0:
        effective = precip_mm * 0.20
    elif precip_mm < 10.0:
        effective = precip_mm * 0.60
    else:
        effective = precip_mm * 0.90
    relief = min(1.5, effective / 5.0)

    return max(0.0, base + et0_adj - relief)


def _apply_drought_modifiers(days: list[dict], region_name: str) -> list[dict]:
    """Apply phenological weighting, persistence penalty, and regional overrides.

    Phenology (critical growth windows amplify sensitivity):
        Aug–Oct (flowering + fruit fill): ×1.2
        Jan–Feb (grand expansion):        ×1.1
        All other months:                 ×1.0

    Persistence penalty (cumulative stress, not one dry day):
        If 10+ of the 14 forecast days have score > 1.0 → add +0.5 to ALL days

    Robusta VPD override (Espírito Santo — Conilon stresses at lower VPD):
        If VPD > 1.5 kPa → floor score at 2.0 (minimum M)

    Classifies drought_risk: ≥ 2.5 → H, ≥ 1.5 → M, ≥ 0.5 → L, else → —
    """
    PHENO = {8: 1.2, 9: 1.2, 10: 1.2, 1: 1.1, 2: 1.1}

    # Step 1 — Apply phenology only (no Robusta override yet)
    for day in days:
        month = int(day["date"][5:7])
        score = day.get("_drought_score_raw", 0.0)
        score *= PHENO.get(month, 1.0)
        day["_drought_score"] = score

    # Step 2 — Persistence penalty (based on natural formula, before override)
    stressed = sum(1 for d in days if d["_drought_score"] > 1.0)
    penalty = 0.5 if len(days) >= 14 and stressed >= 10 else 0.0

    # Step 3 — Final classify, applying Robusta override AFTER persistence
    for day in days:
        score = day["_drought_score"] + penalty
        # Robusta override: floor at 1.5 (bottom of M) when VPD > 1.5 kPa
        if region_name == "Espírito Santo" and day.get("vpd", 0.0) > 1.5:
            score = max(score, 1.5)
        if score >= 2.5:
            day["drought_risk"] = "H"
        elif score >= 1.5:
            day["drought_risk"] = "M"
        elif score >= 0.5:
            day["drought_risk"] = "L"
        else:
            day["drought_risk"] = "-"
        day.pop("_drought_score_raw", None)
        day.pop("_drought_score", None)

    return days


def _aggregate_hourly_to_daily(hourly: dict, region_name: str = "") -> list[dict]:
    """Aggregate Open-Meteo hourly data into daily dicts.

    Expected hourly keys:
        time, temperature_2m, dew_point_2m, cloud_cover, wind_speed_10m,
        precipitation, precipitation_probability, et0_fao_evapotranspiration,
        soil_moisture_0_to_1cm, soil_moisture_1_to_3cm, soil_moisture_3_to_9cm,
        soil_moisture_9_to_27cm, soil_moisture_27_to_81cm

    Frost risk uses conditions at the hour of minimum temperature.
    Drought risk uses the institutional agronomic model (RWC base + ET0 + rain
    relief), with phenology, persistence, and Robusta overrides applied after
    all 14 days are computed.

    Returns a list of dicts, one per calendar day.
    """
    times      = hourly["time"]
    temp       = hourly["temperature_2m"]
    dew        = hourly["dew_point_2m"]
    cloud      = hourly["cloud_cover"]
    wind       = hourly["wind_speed_10m"]
    precip_prob_arr = hourly.get("precipitation_probability", [])
    precip_mm_arr   = hourly.get("precipitation", [])
    et0_arr         = hourly.get("et0_fao_evapotranspiration", [])
    sm0 = hourly.get("soil_moisture_0_to_1cm",  [])
    sm1 = hourly.get("soil_moisture_1_to_3cm",  [])
    sm2 = hourly.get("soil_moisture_3_to_9cm",  [])
    sm3 = hourly.get("soil_moisture_9_to_27cm", [])
    sm4 = hourly.get("soil_moisture_27_to_81cm",[])

    def _safe(arr: list, i: int):
        return arr[i] if i < len(arr) else None

    # Group indices by calendar date
    day_indices: dict[str, list[int]] = defaultdict(list)
    for i, ts in enumerate(times):
        day_indices[ts[:10]].append(i)

    days = []
    for day_str in sorted(day_indices):
        idxs = day_indices[day_str]

        temps_day  = [v for v in (temp[i]  for i in idxs) if v is not None]
        dew_day    = [v for v in (dew[i]   for i in idxs) if v is not None]
        cloud_day  = [v for v in (cloud[i] for i in idxs) if v is not None]
        wind_day   = [v for v in (wind[i]  for i in idxs) if v is not None]
        pp_day     = [v for v in (_safe(precip_prob_arr, i) for i in idxs) if v is not None]
        pm_day     = [v for v in (_safe(precip_mm_arr, i)  for i in idxs) if v is not None]
        et0_day    = [v for v in (_safe(et0_arr, i)        for i in idxs) if v is not None]
        sm0_day    = [v for v in (_safe(sm0, i) for i in idxs) if v is not None]
        sm1_day    = [v for v in (_safe(sm1, i) for i in idxs) if v is not None]
        sm2_day    = [v for v in (_safe(sm2, i) for i in idxs) if v is not None]
        sm3_day    = [v for v in (_safe(sm3, i) for i in idxs) if v is not None]
        sm4_day    = [v for v in (_safe(sm4, i) for i in idxs) if v is not None]

        if not temps_day:
            continue

        min_t = min(temps_day)
        max_t = max(temps_day)

        # Root-zone soil moisture (weighted 3-band)
        surface_vals = sm0_day + sm1_day + sm2_day
        mean_surface = sum(surface_vals) / len(surface_vals) if surface_vals else 0.22
        mean_sm3     = sum(sm3_day) / len(sm3_day) if sm3_day else 0.22
        mean_sm4     = sum(sm4_day) / len(sm4_day) if sm4_day else 0.24
        root_zone    = _root_zone_moisture(mean_surface, mean_sm3, mean_sm4)

        # Frost: conditions at hour of minimum temperature
        min_idx      = min(idxs, key=lambda i: temp[i] if temp[i] is not None else float("inf"))
        cloud_at_min = cloud[min_idx] if cloud[min_idx] is not None else (
            sum(cloud_day) / len(cloud_day) if cloud_day else 0.0
        )
        wind_at_min  = wind[min_idx] if wind[min_idx] is not None else (
            max(wind_day) if wind_day else 0.0
        )
        dew_at_min   = dew[min_idx] if dew[min_idx] is not None else (
            sum(dew_day) / len(dew_day) if dew_day else 0.0
        )

        # VPD from daily max temp + mean dew point
        mean_dew   = sum(dew_day) / len(dew_day) if dew_day else 0.0
        vpd        = _vpd_kpa(max_t, mean_dew)

        # Drought inputs — actual mm values for the formula
        daily_precip_mm = sum(pm_day)  if pm_day  else 0.0
        daily_et0_mm    = sum(et0_day) if et0_day else 0.0
        pp_max          = max(pp_day)  if pp_day  else 0.0

        raw_score = _drought_score(daily_precip_mm, daily_et0_mm, root_zone)

        days.append({
            "date":              day_str,
            "min_temp":          round(min_t, 1),
            "max_temp":          round(max_t, 1),
            "dew_point":         round(mean_dew, 1),
            "cloud_cover":       round(sum(cloud_day) / len(cloud_day), 1) if cloud_day else 0.0,
            "wind_speed":        round(max(wind_day), 1) if wind_day else 0.0,
            "precip_prob":       round(pp_max, 1),
            "precip_mm":         round(daily_precip_mm, 1),
            "et0_mm":            round(daily_et0_mm, 2),
            "soil_moisture":     round(root_zone, 3),
            "vpd":               round(vpd, 2),
            "frost_risk":        _frost_risk(min_t, cloud_at_min, wind_at_min, dew_at_min),
            "drought_risk":      "-",            # filled by _apply_drought_modifiers
            "_drought_score_raw": raw_score,
        })

    # Apply phenology, persistence, and regional overrides to all 14 days together
    _apply_drought_modifiers(days, region_name)

    return days


def _parse_oni_text(text: str) -> list[dict]:
    """
    Parse the NOAA CPC ONI plain-text file.

    Format:
        SEAS YR TOTAL ANOM
        DJF 2024 0.5 1.2
        ...

    All rows are historical 3-month running means. NOAA publishes with ~1-2 month
    lag, so the most recent 3 seasons are marked preliminary (not yet finalized).
    Returns the last 18 rows, with the final 3 flagged preliminary=True.
    """
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("SEAS"):
            continue
        parts = line.split()
        if len(parts) < 4:
            continue
        seas, yr, _total, anom = parts[0], parts[1], parts[2], parts[3]
        if seas not in _SEASON_MONTH:
            continue
        month_num = _SEASON_MONTH[seas]
        year_2d   = int(yr) % 100
        label     = f"{_MONTH_ABBR[month_num]}-{year_2d:02d}"
        rows.append({"month": label, "value": float(anom)})

    if not rows:
        return []

    N_PRELIMINARY = 3
    tail = rows[-18:]  # show 18 months of history
    cutoff = len(tail) - N_PRELIMINARY
    result = []
    for i, r in enumerate(tail):
        entry: dict = {"month": r["month"], "value": r["value"]}
        if i >= cutoff:
            entry["preliminary"] = True
        result.append(entry)
    return result


# ---------------------------------------------------------------------------
# DB-writing functions
# ---------------------------------------------------------------------------

def _scrape_weather(db) -> None:
    """Fetch 14-day hourly forecast for each region, aggregate to daily, upsert."""
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
    from models import WeatherSnapshot
    from sqlalchemy import delete

    for region in REGIONS:
        url = OPEN_METEO_URL.format(lat=region["lat"], lon=region["lon"])
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            hourly = data.get("hourly", {})
            daily_data = _aggregate_hourly_to_daily(hourly, region_name=region["name"])

            db.execute(delete(WeatherSnapshot).where(WeatherSnapshot.region == region["name"]))
            db.add(WeatherSnapshot(region=region["name"], daily_data=daily_data, scraped_at=datetime.utcnow()))
            db.commit()
            print(f"[farmer_economics] weather OK: {region['name']} ({len(daily_data)} days)")
        except Exception as e:
            db.rollback()
            print(f"[farmer_economics] weather FAILED for {region['name']}: {e}")


def _scrape_enso(db) -> None:
    """Fetch NOAA ONI text, parse, store in NewsItem."""
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
    from models import NewsItem
    from sqlalchemy import delete

    try:
        resp = requests.get(NOAA_ONI_URL, timeout=30)
        resp.raise_for_status()
        oni_history = _parse_oni_text(resp.text)

        db.execute(delete(NewsItem).where(NewsItem.source == "NOAA CPC"))
        db.add(NewsItem(
            title=f"ENSO ONI – {date.today():%Y-%m}",
            source="NOAA CPC",
            category="supply",
            tags=["enso", "supply", "brazil"],
            meta=json.dumps({"oni_history": oni_history}),
            pub_date=datetime.utcnow(),
        ))
        db.commit()
        print(f"[farmer_economics] ENSO OK ({len(oni_history)} rows)")
    except Exception as e:
        db.rollback()
        print(f"[farmer_economics] ENSO FAILED: {e}")


IRI_FORECAST_URL = "https://iri.columbia.edu/our-expertise/climate/forecasts/enso/current/"

_SEASON_ORDER = ["DJF","JFM","FMA","MAM","AMJ","MJJ","JJA","JAS","ASO","SON","OND","NDJ"]


def _scrape_enso_forecast(db) -> None:
    """Fetch IRI/CPC ENSO probability table and append to the NOAA ONI item.

    Table layout (column-oriented):
        Season | La Nina | Neutral | El Nino
        MAM    |   0     |   91    |    9
        AMJ    |   0     |   53    |   47
        ...
    """
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
    from models import NewsItem
    from bs4 import BeautifulSoup

    try:
        resp = requests.get(IRI_FORECAST_URL, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        forecast: list[dict] = []
        for table in soup.find_all("table"):
            all_text = table.get_text()
            if "La Nina" not in all_text and "La Ni\xf1a" not in all_text:
                continue

            rows = table.find_all("tr")
            if not rows:
                continue

            # Header row: Season | La Nina | Neutral | El Nino
            headers = [th.get_text(strip=True).lower() for th in rows[0].find_all(["th", "td"])]
            try:
                col_season  = headers.index("season")
                col_lanina  = next(i for i, h in enumerate(headers) if "la" in h and ("nina" in h or "niña" in h))
                col_neutral = next(i for i, h in enumerate(headers) if "neutral" in h)
                col_elnino  = next(i for i, h in enumerate(headers) if "el" in h and ("nino" in h or "niño" in h))
            except (ValueError, StopIteration):
                continue

            # Data rows
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if len(cells) <= max(col_season, col_lanina, col_neutral, col_elnino):
                    continue
                season = cells[col_season].get_text(strip=True)
                if len(season) != 3 or not season.isupper():
                    continue

                def _int(cell):
                    try:
                        return int(cell.get_text(strip=True).replace("%", ""))
                    except ValueError:
                        return None

                forecast.append({
                    "season":  season,
                    "la_nina": _int(cells[col_lanina]),
                    "neutral": _int(cells[col_neutral]),
                    "el_nino": _int(cells[col_elnino]),
                })

            if forecast:
                break

        if not forecast:
            print("[farmer_economics] ENSO forecast: probability table not found")
            return

        # Append to existing ONI NewsItem meta
        enso_item = (
            db.query(NewsItem)
            .filter(NewsItem.source == "NOAA CPC")
            .order_by(NewsItem.pub_date.desc())
            .first()
        )
        if enso_item:
            meta = json.loads(enso_item.meta or "{}")
            meta["oni_forecast"] = forecast
            enso_item.meta = json.dumps(meta)
            db.commit()
            print(f"[farmer_economics] ENSO forecast OK ({len(forecast)} seasons): "
                  f"{[f['season'] for f in forecast]}")
        else:
            print("[farmer_economics] ENSO forecast: no ONI item to attach to")
    except Exception as e:
        db.rollback()
        print(f"[farmer_economics] ENSO forecast FAILED: {e}")


def _parse_world_bank_excel(content: bytes) -> dict:
    """Parse World Bank CMO Pink Sheet Excel.

    The sheet is wide-format:
      - Row 4 (0-indexed): commodity header labels across columns
      - Column A (index 0): month strings like "2024M12"
      - Data at intersections

    Returns {"urea_monthly": [...7 vals...], "dap_monthly": [...], "kcl_monthly": [...]}.
    """
    import re
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    if "Monthly Prices" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Monthly Prices' not found. Available: {wb.sheetnames}")

    ws = wb["Monthly Prices"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Step 1: find the header row — the row that contains "DAP" somewhere
    header_row_idx = None
    for i, row in enumerate(rows):
        for cell in row:
            if isinstance(cell, str) and cell.strip().upper() == "DAP":
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return {"urea_monthly": [], "dap_monthly": [], "kcl_monthly": []}

    header = rows[header_row_idx]

    # Step 2: find column indices for Urea, DAP, KCl
    col_map: dict[str, int] = {}
    for i, cell in enumerate(header):
        if not isinstance(cell, str):
            continue
        s = cell.strip().lower()
        if s == "urea" and "urea" not in col_map:
            col_map["urea"] = i
        elif s == "dap" and "dap" not in col_map:
            col_map["dap"] = i
        elif s.startswith("potassium chloride") and "kcl" not in col_map:
            col_map["kcl"] = i

    if not col_map:
        return {"urea_monthly": [], "dap_monthly": [], "kcl_monthly": []}

    # Step 3: collect data rows — column A matches "YYYYMxx"
    month_re = re.compile(r"^\d{4}M\d{2}$")
    data_rows = [
        row for row in rows[header_row_idx + 1:]
        if row and isinstance(row[0], str) and month_re.match(row[0].strip())
    ]

    # Step 4: extract last 7 numeric values per commodity
    def _extract(key: str) -> list[float]:
        if key not in col_map:
            return []
        col = col_map[key]
        vals = []
        for row in data_rows:
            v = row[col] if col < len(row) else None
            if isinstance(v, (int, float)) and v is not None:
                vals.append(float(v))
        return vals[-7:] if len(vals) >= 7 else vals

    # Step 5: find last month string that has data for at least one commodity
    last_data_month = None
    _MONTH_MAP = {
        "01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
        "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec",
    }
    for row in reversed(data_rows):
        if not row or not isinstance(row[0], str):
            continue
        has_data = any(
            isinstance(row[col_map[k]], (int, float))
            for k in col_map
            if col_map[k] < len(row)
        )
        if has_data:
            raw = row[0].strip()  # e.g. "2026M02"
            yr, mo = raw[:4], raw[5:7]
            last_data_month = f"{_MONTH_MAP.get(mo, mo)}-{yr}"
            break

    return {
        "urea_monthly":    _extract("urea"),
        "dap_monthly":     _extract("dap"),
        "kcl_monthly":     _extract("kcl"),
        "last_data_month": last_data_month,
    }


def _find_world_bank_excel_url() -> str:
    """Scrape World Bank commodity-markets page to find the current Monthly Excel URL.

    The document ID in the URL rotates with every file update, so we discover
    it dynamically instead of hard-coding it. Falls back to WORLD_BANK_FALLBACK_URL
    if the page is unreachable or no link is found.
    """
    import re as _re
    try:
        resp = requests.get(WORLD_BANK_COMMODITY_PAGE, timeout=30,
                            headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        # Find any thedocs.worldbank.org link containing CMO-Historical-Data-Monthly.xlsx
        match = _re.search(
            r'https://thedocs\.worldbank\.org/en/doc/[^"\']+CMO-Historical-Data-Monthly\.xlsx',
            resp.text,
        )
        if match:
            url = match.group(0)
            print(f"[farmer_economics] World Bank Excel URL discovered: {url}")
            return url
    except Exception as e:
        print(f"[farmer_economics] Could not discover World Bank URL: {e}")
    print(f"[farmer_economics] Falling back to: {WORLD_BANK_FALLBACK_URL}")
    return WORLD_BANK_FALLBACK_URL


def _scrape_fertilizer_prices(db) -> None:
    """Download World Bank Excel, extract last 7 monthly values for urea/DAP/KCl."""
    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
    from models import NewsItem
    from sqlalchemy import delete

    try:
        url = _find_world_bank_excel_url()
        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
        parsed = _parse_world_bank_excel(resp.content)

        db.execute(delete(NewsItem).where(NewsItem.source == "World Bank"))
        db.add(NewsItem(
            title=f"Fertilizer Prices – {date.today():%Y-%m}",
            source="World Bank",
            category="supply",
            tags=["fertilizer", "price", "supply", "brazil"],
            meta=json.dumps(parsed),
            pub_date=datetime.utcnow(),
        ))
        db.commit()
        print(f"[farmer_economics] fertilizer OK — keys: {[k for k, v in parsed.items() if v]}")
    except Exception as e:
        db.rollback()
        print(f"[farmer_economics] fertilizer FAILED: {e}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

async def run(page, db) -> None:
    """
    Entry point called from main.py.
    `page` is a Playwright page (unused here; kept for signature consistency).
    """
    _scrape_weather(db)
    _scrape_enso(db)
    _scrape_enso_forecast(db)
    _scrape_fertilizer_prices(db)
