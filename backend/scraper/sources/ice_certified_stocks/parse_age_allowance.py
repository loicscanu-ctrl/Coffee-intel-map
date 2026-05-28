"""Robusta age-allowance report (.xlsx — monthly, end-of-month publish).

Source: …/aged_allowance_stock_report/Robusta_Coffee_Age_Allowance_YYYYMMDD.xlsx

Two sheets ("Valid Stock Table (2)" + "Transition Stock Table (2)") with the
same schema: 30 monthly buckets (Months Since Graded = 1..30) × 13 ports
(AMS/ANT/BAR/BRE/FEL/GEN/HAM/LEH/LON/NYK/NOR/ROT/TRI) in metric tonnes, plus a
Grand Total (MT) column and a Total Age Allowance ($/MT) column that ramps
$5/month from month 13 (first 12 months are the grace period — "n/a").
"""
from __future__ import annotations

import io

import openpyxl


def _coerce_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _coerce_allow(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, str) and v.strip().lower() in {"n/a", "na", "-", ""}:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_sheet(sheet) -> dict:
    rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]

    # Locate the header row (the one whose first non-empty cell is "Months Since Graded").
    header_idx: int | None = None
    for i, r in enumerate(rows):
        if any(isinstance(c, str) and c.strip() == "Months Since Graded" for c in r):
            header_idx = i; break
    if header_idx is None:
        return {"ports": [], "buckets": [], "totals_by_port": {}, "grand_total_mt": 0, "report_date": None}

    header = rows[header_idx]
    months_col = grand_col = allow_col = None
    port_cols: list[tuple[int, str]] = []
    for c, h in enumerate(header):
        if not isinstance(h, str):
            continue
        hs = h.strip()
        if hs == "Months Since Graded":
            months_col = c
        elif hs == "Grand Total (MT)":
            grand_col = c
        elif "Total Age Allowance" in hs:
            allow_col = c
        elif len(hs) == 3 and hs.isupper():
            port_cols.append((c, hs))

    buckets: list[dict] = []
    totals_by_port: dict[str, int] = {}
    grand_total = 0
    for r in rows[header_idx + 1:]:
        if not r or months_col is None:
            continue
        mb = r[months_col]
        if isinstance(mb, str) and mb.strip().upper() == "TOTAL":
            totals_by_port = {p: _coerce_int(r[c]) for c, p in port_cols}
            grand_total = _coerce_int(r[grand_col]) if grand_col is not None else sum(totals_by_port.values())
            continue
        try:
            month_n = int(mb)
        except (TypeError, ValueError):
            continue
        buckets.append({
            "months_since_graded":      month_n,
            "by_port":                  {p: _coerce_int(r[c]) for c, p in port_cols},
            "grand_total_mt":           _coerce_int(r[grand_col]) if grand_col is not None else None,
            "age_allowance_usd_per_mt": _coerce_allow(r[allow_col]) if allow_col is not None else None,
        })

    # Report date is a datetime cell somewhere in the first ~5 rows.
    report_date_iso: str | None = None
    for r in rows[:5]:
        for c in r:
            if hasattr(c, "isoformat"):
                try:
                    report_date_iso = c.date().isoformat()
                except Exception:
                    pass
                break
        if report_date_iso:
            break

    return {
        "report_date":    report_date_iso,
        "ports":          [p for _, p in port_cols],
        "buckets":        buckets,
        "totals_by_port": totals_by_port,
        "grand_total_mt": grand_total,
    }


def parse_age_allowance_xlsx(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: dict[str, dict | None] = {"valid": None, "transition": None, "report_date": None}
    for sname in wb.sheetnames:
        ln = sname.lower()
        parsed = _parse_sheet(wb[sname])
        if "valid" in ln:
            out["valid"] = parsed
        elif "transition" in ln:
            out["transition"] = parsed
        # Use the first parsed date as the report date.
        if out["report_date"] is None and parsed.get("report_date"):
            out["report_date"] = parsed["report_date"]
    return out
